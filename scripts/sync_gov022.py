#!/usr/bin/env python3
"""ATHENA-GOV-022: AI Intelligence V3 Research documentation into the XLSX (matches Markdown).
DEC-AI-RESEARCH-001..007, V3 backlog/epics/bugs, CHANGELOG row. Idempotent; backup."""
import openpyxl, shutil, datetime, os

P = os.path.join('docs', 'PIA_MASTER_BACKLOG_SOURCE_OF_TRUTH.xlsx')
BR = 'feat/pia-v3-foundation-integration'


def ids(ws, col='A'):
    return {str(c.value).strip() for c in ws[col] if c.value not in (None, '')}


def main():
    bak = P + '.bak.' + datetime.datetime.now(datetime.UTC).strftime('%Y%m%d%H%M%S')
    shutil.copy2(P, bak); print('backup', bak)
    wb = openpyxl.load_workbook(P)

    ad = wb['Architecture Decisions']; have_ad = ids(ad)
    decisions = [
        ['DEC-AI-RESEARCH-001', 'AI Intelligence', 'Research tab is thesis-only; it must not show Buy/Hold/Sell recommendation logic.', 'LOCKED', 'Research explains the thesis; verdict lives in Overview.'],
        ['DEC-AI-RESEARCH-002', 'AI Intelligence', 'Ownership split: Overview owns verdict/action; Portfolio owns position action; Research owns thesis/deep analysis.', 'LOCKED', 'Clear separation of concerns.'],
        ['DEC-AI-RESEARCH-003', 'AI Intelligence', 'No dummy data allowed; missing provider data must be shown as missing/partial or hidden.', 'LOCKED', 'Auditable, trustworthy research.'],
        ['DEC-AI-RESEARCH-004', 'AI Intelligence', 'Competitive Comparison must only render with real backend-supported peer data (shouldRender=true).', 'LOCKED', 'No hardcoded/dummy/fallback peers.'],
        ['DEC-AI-RESEARCH-005', 'AI Intelligence', 'Research V2 approved mock is docs/mocks/ai-intelligence/APPROVED/research-approved.png (implementation source of truth). WARNING: asset currently MISSING from repo — reference broken; Design Lock Package INVALID per DESIGN-LOCK-002 until the approved image is committed.', 'LOCKED (ASSET MISSING)', 'Asset must be archived/committed to validate the lock.'],
        ['DEC-AI-RESEARCH-006', 'AI Intelligence', 'Accordion arrows: collapsed = down arrow, expanded = up arrow (down encourages expansion).', 'LOCKED', 'Affordance for deep-dive expansion.'],
        ['DEC-AI-RESEARCH-007', 'AI Intelligence', 'Research customization locked: show/hide sections, drag reorder, text size S/M/L/XL, default expanded state, persist preferences.', 'LOCKED', 'Institutional-style configurable research.'],
    ]
    n_ad = 0
    for r in decisions:
        if r[0] in have_ad:
            continue
        ad.append(r); n_ad += 1
    print('Architecture Decisions appended', n_ad)

    bl = wb['Backlog']; have_b = ids(bl)
    rows = [
        ['HERMES-AI-V3-001', 'AI Intelligence', 'AI Intelligence V3 Research', 'Research Backend Gap Analysis (coverage matrix for 9 sections, data-source mapping, proposed contract, provider gaps, thesis-only constraint)', 'Artifact: docs/HERMES-AI-V3-001_RESEARCH_BACKEND_GAP_ANALYSIS.md', 'P1', 'COMPLETED', 'HERMES', BR, 'PASS', 'ACCEPTED', 'ACCEPTED', 'Backend-only audit'],
        ['HERMES-AI-V3-002', 'AI Intelligence', 'AI Intelligence V3 Research', 'Research Endpoint V1: ai_research.py + GET /api/intelligence/{symbol}/research; thesis-only; explicit null/status placeholders', 'p50 9.93ms / p95 11.86ms; contract validator + example payload', 'P1', 'COMPLETED', 'HERMES', BR, 'PASS', 'ACCEPTED', 'ACCEPTED', 'No Buy/Hold/Sell; no portfolio action'],
        ['HERMES-AI-V3-003', 'AI Intelligence', 'AI Intelligence V3 Research', 'Research Provenance & Real Data Upgrade: schema V3.0, ResearchMetric + section provenance, competitiveComparison shouldRender=false when no provider; auditable null placeholders', 'p50 12.12ms / p95 17.29ms (well under 500/1000ms)', 'P1', 'COMPLETED', 'HERMES', BR, 'PASS', 'ACCEPTED', 'ACCEPTED', 'No dummy/fallback peers'],
        ['CR-AI-V3-UI-001', 'AI Intelligence', 'AI Intelligence V3', 'Overview/Compact/Expanded Hero corrections (C1-3, E1-7, S1-2: remove case badge, hero sizing, risk label, section header standard, breakpoints)', 'Corrected deviation from approved mock', 'P1', 'CLOSED', 'ARTEMIS', BR, 'PASS', 'PASS', 'CLOSED', 'Commit 89bad3a'],
        ['ARTEMIS-AI-V3-RESEARCH-003', 'AI Intelligence', 'AI Intelligence V3 Research', 'Research V2 final implementation (Research tab full impl + proxy route)', 'GOVERNANCE GAP: approved mock research-approved.png is MISSING; Design Lock Package invalid (DESIGN-LOCK-002)', 'CRITICAL', 'IMPLEMENTED (LOCK INVALID)', 'ARTEMIS', BR, 'PASS', 'BLOCKED - missing approved mock', 'PENDING', 'Commits 8657868, b056bc1; needs approved-asset commit + UAT screenshots'],
        ['EPIC-AI-RESEARCH-V2', 'AI Intelligence', 'AI Intelligence V3 Research', 'Institutional Research Memo: 11 sections, customization, provenance, real backend data only', 'Dep: HERMES-AI-V3-003 done; ARTEMIS-AI-V3-RESEARCH-003', 'P1', 'IN PROGRESS', 'ARTEMIS / HERMES', BR, '', 'PENDING', 'PENDING', 'Success: 390px matches mock; no dummy data; <1s load; customizable'],
        ['EPIC-AI-PROVENANCE', 'AI Intelligence', 'AI Intelligence V3 Research', 'Data Provenance & Trust Layer: section + metric provenance, source/lastUpdated/refresh/confidence/dataType/calcMethod', 'Backend V1 done in HERMES-AI-V3-003; frontend rendering pending Research V2', 'P1', 'BACKEND COMPLETE', 'HERMES / ARTEMIS', BR, '', '', '', 'Frontend pending'],
        ['EPIC-AI-COMPETITIVE-COMPARISON', 'AI Intelligence', 'AI Intelligence V3 Research', 'Competitive Comparison Engine: real peer-based comparison; no hardcoded/dummy/fallback peers; shouldRender=false if provider unavailable', 'Backend placeholder only; future peer-selection provider', 'P2', 'BACKLOG', 'HERMES', BR, '', '', '', 'Future provider work'],
        ['BUG-HERMES-AI-007-AMD-MATERIAL-NEWS', 'AI Intelligence', 'AI Intelligence V2.5', 'Older HERMES-AI-007 validator fails on AMD (material news impact missing); not caused by Research V1/V3; does not block Research endpoint', 'V2.5 dynamic news acceptance path', 'P2', 'OPEN', 'HERMES', BR, '', '', '', 'Found during HERMES-AI-V3-002 regression'],
        ['BUG-AI-RESEARCH-COMPETITIVE-DATA-MISSING', 'AI Intelligence', 'AI Intelligence V3 Research', 'Competitive Comparison cannot render real peers (no peer/metrics provider); correctly returns status=missing, shouldRender=false', 'Known gap; not blocking Research V2 UI when hidden', 'P1', 'KNOWN GAP', 'HERMES', BR, '', '', '', 'Needs peer provider'],
        ['BUG-AI-RESEARCH-PROVIDER-GAPS', 'AI Intelligence', 'AI Intelligence V3 Research', 'Missing providers: normalized financials, TAM/segment, revenue-vs-estimate, guidance, institutional ownership trends, fund sentiment, DCF; explicit null/status placeholders', 'Known gap; explicit auditable placeholders', 'P2', 'KNOWN GAP', 'HERMES', BR, '', '', '', 'Provider integration future'],
        ['GOV-022-RESEARCH-MOCK-MISSING', 'Governance', 'Mock Compliance', 'Approved Research mock missing: research-approved.png absent (and typo research-aproved.png + drafts no longer present). Archive the approved image under docs/mocks/ai-intelligence/APPROVED/ and add RESEARCH_DESIGN_SPEC.md to validate the Research V2 Design Lock.', 'Blocks UAT/closure of ARTEMIS-AI-V3-RESEARCH-003', 'P0', 'OPEN', 'ATHENA / PO', BR, '', '', '', 'DEC-AI-RESEARCH-005 reference currently broken'],
    ]
    n_b = 0
    for r in rows:
        if r[0] in have_b:
            continue
        bl.append(r); n_b += 1
    print('Backlog appended', n_b)

    ch = wb['CHANGELOG']
    if ch.max_row == 1 and (ch['A1'].value in (None, '')):
        ch['A1'], ch['B1'], ch['C1'], ch['D1'] = 'Date', 'Commit', 'Author', 'Message'
    have_c = {str(c.value).strip() for c in ch['B'] if c.value}
    crows = [('2026-06-22', 'GOV-022', 'ATHENA', 'AI Intelligence V3 Research documentation: HERMES-AI-V3-001/002/003 COMPLETE; CR-AI-V3-UI-001 CLOSED (89bad3a); ARTEMIS-AI-V3-RESEARCH-003 IMPLEMENTED but approved mock MISSING; DEC-AI-RESEARCH-001..007 LOCKED; epics/bugs added')]
    n_ch = 0
    for d, c, a, m in crows:
        if c not in have_c:
            ch.append([d, c, a, m]); n_ch += 1
    print('CHANGELOG appended', n_ch)

    wb.save(P); print('SAVED', P)


if __name__ == '__main__':
    main()
