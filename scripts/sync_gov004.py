#!/usr/bin/env python3
"""PIA-GOV-004 registration into the XLSX (matches the Markdown).
Adds DEC-GOV-004 (Architecture Decisions, LOCKED), GOV-004-REMEDIATION (Backlog),
and a CHANGELOG sheet row. Idempotent; timestamped backup."""
import openpyxl, shutil, datetime, os

P = os.path.join('docs', 'PIA_MASTER_BACKLOG_SOURCE_OF_TRUTH.xlsx')
BR = 'feat/pia-v3-foundation-integration'


def ids(ws, col='A'):
    return {str(c.value).strip() for c in ws[col] if c.value not in (None, '')}


def main():
    bak = P + '.bak.' + datetime.datetime.now(datetime.UTC).strftime('%Y%m%d%H%M%S')
    shutil.copy2(P, bak); print('backup', bak)
    wb = openpyxl.load_workbook(P)

    ad = wb['Architecture Decisions']; have = ids(ad)
    dec = ['DEC-GOV-004', 'Process Governance',
           'Approved Mock Preservation & Design Lock Traceability: every Design Lock must archive the approved mock under docs/mocks/<feature>/APPROVED_<feature>_v<version>.png and COMMIT it to git BEFORE implementation starts; record the approved-mock path in the backlog item, UAT ticket, and Design Lock notes. Process: Requirement -> UX Mockup -> Design Review -> Design Lock -> SAVE approved mock -> COMMIT approved mock -> Implementation -> UAT. Every UAT report must include Approved Mock <path>, Design Lock Commit <id>, Implementation Commit <id>. Non-compliance: implementation started without an archived approved mock is a governance violation and is blocked until the mock is committed.',
           'LOCKED',
           'Analyst Targets drifted because the approved mock was not preserved as a repository source of truth (PIA-GOV-004).']
    n_ad = 0
    if dec[0] not in have:
        ad.append(dec); n_ad += 1
    print('Architecture Decisions appended', n_ad)

    bl = wb['Backlog']; have_b = ids(bl)
    item = ['GOV-004-REMEDIATION', 'Governance', 'Approved Mock Compliance',
            'Normalize existing approved mocks to APPROVED_<feature>_v<version>.png; consolidate docs/mocks vs docs/design-system/mocks; resolve "AI Intelligence/mock v1.png" vs locked AI Intelligence V2 version; backfill traceability triples (Approved Mock / Design Lock Commit / Implementation Commit).',
            'Non-compliant assets: docs/mocks/AI Intelligence/mock v1.png; docs/mocks/analyst-targets/Approved_mobile_mock_analyst_target.jpg; docs/mocks/stock-intelligence/stock-intelligence-v1-approved.png; docs/mocks/watchlists/watchlists-mobile-v1-approved.md',
            'P1', 'OPEN', 'ATHENA', BR, '', '', '', 'Created by PIA-GOV-004 (DEC-GOV-004 LOCKED)']
    n_b = 0
    if item[0] not in have_b:
        bl.append(item); n_b += 1
    print('Backlog appended', n_b)

    ch = wb['CHANGELOG']
    if ch.max_row == 1 and (ch['A1'].value in (None, '')):
        ch['A1'], ch['B1'], ch['C1'], ch['D1'] = 'Date', 'Commit', 'Author', 'Message'
    have_c = {str(c.value).strip() for c in ch['B'] if c.value}
    rows = [('2026-06-11', 'GOV-004', 'ATHENA',
             'PIA-GOV-004 Approved Mock Preservation & Design Lock Traceability LOCKED (DEC-GOV-004); compliance audit + GOV-004-REMEDIATION created')]
    n_ch = 0
    for d, c, a, m in rows:
        if c not in have_c:
            ch.append([d, c, a, m]); n_ch += 1
    print('CHANGELOG appended', n_ch)

    wb.save(P); print('SAVED', P)


if __name__ == '__main__':
    main()
