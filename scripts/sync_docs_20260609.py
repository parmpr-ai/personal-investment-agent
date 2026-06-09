#!/usr/bin/env python3
"""Documentation sync 2026-06-09: append backlog/UAT/decisions/changelog rows
to PIA_MASTER_BACKLOG_SOURCE_OF_TRUTH.xlsx so it matches the Markdown source.
Idempotent: skips rows whose key already exists. Creates a timestamped backup."""
import openpyxl, shutil, datetime, os

P = os.path.join('docs', 'PIA_MASTER_BACKLOG_SOURCE_OF_TRUTH.xlsx')
BR = 'feat/pia-v3-foundation-integration'


def col_ids(ws, col='A'):
    return {str(c.value).strip() for c in ws[col] if c.value not in (None, '')}


def main():
    bak = P + '.bak.' + datetime.datetime.utcnow().strftime('%Y%m%d%H%M%S')
    shutil.copy2(P, bak)
    print('backup', bak)
    wb = openpyxl.load_workbook(P)

    # Backlog: A ID,B Cat,C Epic,D Item,E Desc,F Pri,G Status,H Owner,I Branch,J Tech,K UAT,L PO,M Notes
    bl = wb['Backlog']
    have = col_ids(bl)
    backlog = [
        ['CR-AT-V3', 'Stock Intelligence', 'Analyst Targets V3',
         'V3 layout: Options tab removed; chart only in Chart tab; fixed sticky header; Overview Bull/Base/Bear + target range + consensus + analyst distribution; Analysis subsection (Consensus, Bull/Base/Bear, Distribution, History cards)',
         'Tap Overview card navigates to Analysis > Analyst Targets; history as mobile cards (no tables)',
         'HIGH', 'IMPLEMENTED', 'HERMES', BR, 'PASS', 'READY FOR UAT', 'PENDING',
         'Commits 2b9d1de,5602655,ecbe06d,ac0ca6f through e5736e9; current data source only (no Finnhub/FMP)'],
        ['SPRINT-PORT-DENSITY', 'Portfolio / Cards', 'Portfolio Density Sprint',
         'Cards v2; card customization framework; grid + filters; 2x2 compact IBKR; live price emphasis (color/pop); visual system v2 (logo/hierarchy/2x2); logo ring; portfolio view selector; mobile density pass + persistence validation',
         'Portfolio & Watchlist card visual system', 'HIGH', 'IMPLEMENTED', 'ATHENA', BR, 'PASS', 'READY FOR UAT', 'PENDING',
         'Commits 23bce57,54bf30e,b7c646f,6038934,edca406,5e8daca,72499e9,e5736e9'],
        ['PIA-UX-060', 'Portfolio / Watchlist', 'Card Visuals', 'Card logo still under-weighted as visual anchor',
         'Logo competes with price/ticker; push size/contrast', 'MEDIUM', 'OPEN', 'ATHENA', BR, '', '', '', 'Visual audit 2026-06-09'],
        ['PIA-BUG-032', 'Workspaces', 'Live Widgets', 'Empty workspace preview widgets read as broken-premium',
         'Faded placeholder panels on Portfolio/Watchlist pages', 'MEDIUM', 'OPEN', 'ATHENA / Platform', BR, '', '', '', 'Visual audit 2026-06-09; ties to live-widget conversion'],
        ['PIA-CSS-001', 'Stock Intelligence', 'Header', 'Duplicated/overriding .stock-intel-header CSS; consolidate before V3 fixed header',
         'h2/price/symbol-mark declared multiple times in globals.css', 'MEDIUM', 'OPEN', 'HERMES', BR, '', '', '', 'Visual audit 2026-06-09'],
        ['PIA-UX-061', 'Portfolio', 'View Mode', 'Cards view discoverability: view mode buried in overflow menu',
         'Consider visible 1x1/2x2/3x3 segmented control', 'LOW', 'OPEN', 'PO decision', BR, '', '', '', 'Visual audit 2026-06-09'],
    ]
    n_bl = 0
    for r in backlog:
        if r[0] in have:
            continue
        bl.append(r); n_bl += 1
    print('Backlog appended', n_bl)

    # UAT Log: A UAT ID,B Date,C Area,D Scenario,E Expected,F Result,G PO,H Notes
    ul = wb['UAT Log']
    seen = {str(c.value).strip() for c in ul['A'] if c.value}
    D = '2026-06-09'
    uat = [
        ['UAT-AT-V3', D, 'Stock Intelligence', 'Analyst Targets V3 Overview + Analysis',
         'Options removed; chart only in Chart tab; Overview bull/base/bear + range + consensus + distribution; tap to Analysis history cards',
         'TECH PASS', 'READY FOR PO UAT', 'Current data source; firm-level history limited without Finnhub/FMP'],
        ['UAT-PORT-DENSITY', D, 'Portfolio / Watchlist', 'Card density, logo, price emphasis, view selector, 2x2/3x3',
         'Build passed; / /mobile /setup 200; cards render across 1x1/2x2/3x3', 'TECH PASS', 'READY FOR PO UAT', 'Screenshots captured desktop 1440 + mobile 390'],
        ['UAT-VIS-001', D, 'Portfolio / Watchlist', 'Visual audit: logo sizing', 'Logo a strong anchor', 'FINDING', 'OPEN', 'Logo still small vs price/ticker (PIA-UX-060)'],
        ['UAT-VIS-002', D, 'Workspaces', 'Visual audit: empty preview widgets', 'Premium-feeling cards', 'FINDING', 'OPEN', 'Empty workspace preview panels look broken-premium (PIA-BUG-032)'],
        ['UAT-VIS-003', D, 'Stock Intelligence', 'Visual audit: stock header CSS', 'Single consistent header', 'FINDING', 'OPEN', 'Duplicated .stock-intel-header rules (PIA-CSS-001)'],
        ['UAT-WL-CARRY', D, 'Watchlists', 'Carry-forward watchlist UAT (PIA-WL-008..014)',
         'Column switches; Open Chart target; Add-to-list; AI Coach; add-instrument UX; sorting; columns', 'OPEN', 'OPEN', 'Carried from 2026-06-02 watchlist UAT; still open'],
    ]
    n_ul = 0
    for r in uat:
        if r[0] in seen:
            continue
        ul.append(r); n_ul += 1
    print('UAT Log appended', n_ul)

    # Architecture Decisions: A ID,B Area,C Decision,D Status,E Rationale
    ad = wb['Architecture Decisions']
    have_ad = col_ids(ad)
    adr = [
        ['DEC-DESIGN-LOCK', 'Process Governance',
         'Design Lock: a feature marked DESIGN LOCKED has its layout/IA frozen; implementation must match the locked spec; deviations require re-approval',
         'LOCKED', 'Introduced with the Analyst Targets V3 design-locked spec'],
        ['DEC-NEXT-CACHE', 'Build Governance',
         'Next.js cache rule: on PageNotFoundError during page-data collection (/_not-found, /_document), clear .next then rebuild; never delete .next while a dev/prod server holds it (file lock); avoid concurrent .next access in the shared working tree',
         'LOCKED', 'Recurring build contention during 2026-06 multi-agent sprints'],
    ]
    n_ad = 0
    for r in adr:
        if r[0] in have_ad:
            continue
        ad.append(r); n_ad += 1
    print('Architecture Decisions appended', n_ad)

    # CHANGELOG sheet: A Date,B Commit,C Author,D Message
    ch = wb['CHANGELOG']
    if ch.max_row == 1 and (ch['A1'].value in (None, '')):
        ch['A1'], ch['B1'], ch['C1'], ch['D1'] = 'Date', 'Commit', 'Author', 'Message'
    have_c = {str(c.value).strip() for c in ch['B'] if c.value}
    commits = [
        ('2026-06-03', '02dfcdf', 'ATHENA', 'move market session line to portfolio header'),
        ('2026-06-03', '0778807', 'HERMES/ATHENA', 'portfolio and hero ux polish'),
        ('2026-06-04', 'b7c646f', 'HERMES', 'portfolio grid and filters sprint'),
        ('2026-06-04', '0e36100', 'HERMES', 'move today range into stock hero'),
        ('2026-06-04', '6038934', 'ATHENA', 'improve live price emphasis'),
        ('2026-06-05', '54bf30e', 'HERMES', 'add portfolio card customization framework'),
        ('2026-06-05', '68c0656', 'HERMES', 'compact stock hero range and sparkline layout'),
        ('2026-06-05', '23bce57', 'HERMES', 'portfolio and watchlist cards v2'),
        ('2026-06-06', '44d7fb1', 'HERMES', 'add analyst price targets widget'),
        ('2026-06-06', '2948a0f', 'HERMES', 'refine analyst targets placement and detail view'),
        ('2026-06-07', 'edca406', 'ATHENA', 'polish portfolio watchlist visual system'),
        ('2026-06-07', '803e929', 'HERMES', 'enhance analyst targets detail view'),
        ('2026-06-08', '5e8daca', 'ATHENA', 'UAT CR-PC-001 2x2 card compact IBKR style'),
        ('2026-06-08', '2b9d1de', 'HERMES', 'implement analyst targets v3 layout'),
        ('2026-06-08', '5602655', 'HERMES', 'prioritize analyst target upside badge'),
        ('2026-06-08', 'ecbe06d', 'HERMES', 'refine analyst targets navigation and density'),
        ('2026-06-09', 'ac0ca6f', 'HERMES', 'compact analyst target percentages'),
        ('2026-06-09', '72499e9', 'ATHENA', 'UAT fix pack: empty widgets, logo ring, portfolio view selector'),
        ('2026-06-09', 'e5736e9', 'ATHENA', 'mobile density pass + persistence validation + view selector polish'),
    ]
    n_ch = 0
    for d, c, a, m in commits:
        if c in have_c:
            continue
        ch.append([d, c, a, m]); n_ch += 1
    print('CHANGELOG appended', n_ch)

    wb.save(P)
    print('SAVED', P)


if __name__ == '__main__':
    main()
