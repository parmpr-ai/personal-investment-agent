#!/usr/bin/env python3
"""ATHENA-GOV-021: AI Intelligence V2 governance refresh into the XLSX (matches Markdown).
DEC-AI-009/010/011 (Architecture Decisions), backlog items, CHANGELOG row. Idempotent; backup."""
import openpyxl, shutil, datetime, os

P = os.path.join('docs', 'PIA_MASTER_BACKLOG_SOURCE_OF_TRUTH.xlsx')
BR = 'feat/pia-v3-foundation-integration'


def ids(ws, col='A'):
    return {str(c.value).strip() for c in ws[col] if c.value not in (None, '')}


def main():
    bak = P + '.bak.' + datetime.datetime.now(datetime.UTC).strftime('%Y%m%d%H%M%S')
    shutil.copy2(P, bak); print('backup', bak)
    wb = openpyxl.load_workbook(P)

    ad = wb['Architecture Decisions']; have_ad = ids(ad)
    decisions = [
        ['DEC-AI-009', 'AI Intelligence',
         'Shared Intelligence Data Layer: AI Intelligence SHALL consume data exclusively through the Shared Intelligence Context Layer; direct provider access from widgets is prohibited. Consumers: AI Intelligence, Analyst Targets, Company, Financials, News, Videos.',
         'LOCKED', 'Avoid duplicate fetch logic; consistency; centralized caching + validation.'],
        ['DEC-AI-010', 'AI Intelligence',
         'AI Verdict Separation: AI Verdict (BUY/HOLD/SELL) and Portfolio Recommendation (ADD/HOLD/TRIM/REDUCE/AVOID) are independent systems. Compact shows AI Verdict only; Expanded may display portfolio recommendation.',
         'LOCKED', 'Separate market verdict from portfolio-position action.'],
        ['DEC-AI-011', 'AI Intelligence',
         'Hero System Standardization: all AI Intelligence states use shared hero assets (Neon Wireframe SVG, lattice geometry, institutional/premium look). Rejected: solid-fill, mascot, cartoon, emoji. Compact and Expanded must use identical hero assets.',
         'LOCKED', 'One premium hero visual language across compact + expanded.'],
    ]
    n_ad = 0
    for r in decisions:
        if r[0] in have_ad:
            continue
        ad.append(r); n_ad += 1
    print('Architecture Decisions appended', n_ad)

    bl = wb['Backlog']; have_b = ids(bl)
    rows = [
        ['HERMES-AI-005', 'AI Intelligence', 'AI Intelligence V2',
         'Shared Intelligence Context Layer: context aggregation, context endpoints, validation suite, coverage validation',
         'Accepted', 'P1', 'COMPLETE', 'HERMES', BR, 'PASS', 'ACCEPTED', 'ACCEPTED', 'Backend foundation'],
        ['HERMES-AI-006', 'AI Intelligence', 'AI Intelligence V2',
         'Cache policy, source freshness metadata, frontend contract lock, example payloads, contract validation, frontend lightweight contract mode',
         'Perf warm 6ms compact / 9ms expanded; cold AAPL 2822 NVDA 2446 AMD 1919 TSM 1924 PLTR 1887 ms', 'P1', 'COMPLETE', 'HERMES', BR, 'PASS', 'ACCEPTED', 'ACCEPTED', 'Contract ready for frontend'],
        ['ARTEMIS-AI-011', 'AI Intelligence', 'AI Intelligence V2',
         'AI Intelligence V2 full implementation (Compact + Expanded) + shared hero',
         'Frontend ~80%', 'P1', 'IN PROGRESS', 'ARTEMIS', BR, 'IN PROGRESS', 'PENDING', 'PENDING', 'Commits 4982058,32a7814,8a1b579'],
        ['CR-AI-011', 'AI Intelligence', 'AI Intelligence V2',
         'Visual parity: expanded V2 UX fixes + compact layout overflow; pixel-match approved design',
         'Release blocker', 'P1', 'OPEN', 'ARTEMIS', BR, 'IN PROGRESS', 'PENDING', 'PENDING', 'Commit e67952f; release-blocking'],
        ['HERMES-AI-007', 'AI Intelligence', 'AI Intelligence V2',
         'Parallel Context Hydration: reduce cold symbol load via parallel provider execution. Current 1.9-2.8s; target <1.5s',
         'Performance', 'P2', 'BACKLOG', 'HERMES', BR, '', '', '', 'Cold-load optimization'],
        ['CR-HERMES-006-01', 'AI Intelligence', 'AI Intelligence V2',
         'Contract Versioning: introduce schemaVersion/contractVersion inside frontendPayload',
         'Contract evolution', 'P3', 'BACKLOG', 'HERMES', BR, '', '', '', ''],
    ]
    n_b = 0
    for r in rows:
        if r[0] in have_b:
            continue
        bl.append(r); n_b += 1
    print('Backlog appended', n_b)

    ch = wb['CHANGELOG']
    if ch.max_row == 1 and (ch['A1'].value in (None, '')):
        ch['A1'], ch['B1'], ch['C1'], ch['D1'] = 'Date', 'Commit', 'Author', 'Message'
    have_c = {str(c.value).strip() for c in ch['B'] if c.value}
    crows = [
        ('2026-06-22', 'GOV-021', 'ATHENA',
         'AI Intelligence V2 governance refresh: DEC-AI-009/010/011 LOCKED; HERMES-AI-005/006 COMPLETE; ARTEMIS-AI-011 IN PROGRESS; CR-AI-011 OPEN; HERMES-AI-007 + CR-HERMES-006-01 BACKLOG; RC pending UAT'),
    ]
    n_ch = 0
    for d, c, a, m in crows:
        if c not in have_c:
            ch.append([d, c, a, m]); n_ch += 1
    print('CHANGELOG appended', n_ch)

    wb.save(P); print('SAVED', P)


if __name__ == '__main__':
    main()
