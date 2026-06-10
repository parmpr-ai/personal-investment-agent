#!/usr/bin/env python3
"""AI Intelligence V2 Design Lock registration into the XLSX (matches the Markdown).
Appends DEC-AI-001/002/003 (Architecture Decisions), CR-AI-010 (Backlog),
and a CHANGELOG sheet row. Idempotent; timestamped backup."""
import openpyxl, shutil, datetime, os

P = os.path.join('docs', 'PIA_MASTER_BACKLOG_SOURCE_OF_TRUTH.xlsx')
BR = 'feat/pia-v3-foundation-integration'
SPEC = 'docs/design-system/mocks/stock-workspace/ai-intelligence-widget-v2.md'


def ids(ws, col='A'):
    return {str(c.value).strip() for c in ws[col] if c.value not in (None, '')}


def main():
    bak = P + '.bak.' + datetime.datetime.now(datetime.UTC).strftime('%Y%m%d%H%M%S')
    shutil.copy2(P, bak); print('backup', bak)
    wb = openpyxl.load_workbook(P)

    # Architecture Decisions: A ID,B Area,C Decision,D Status,E Rationale
    ad = wb['Architecture Decisions']; have = ids(ad)
    decisions = [
        ['DEC-AI-001', 'AI Intelligence',
         'KPI Cards: replace KPI rings with KPI cards (Value, Trend Delta, Label, Status, Chevron); full-card tap target; no ring gauges; no flat statistic tiles. Score family (Momentum/Trend/Sentiment, 0-100) vs Directional family (Institutional Flow, Price vs Fair Value) must be visually distinct.',
         'LOCKED', 'Better information density, mobile usability, larger tap targets, improved explainability (AI Intelligence V2)'],
        ['DEC-AI-002', 'AI Intelligence',
         'Single Bottom Sheet Explainability: tap KPI opens one scrollable bottom sheet with Why It Matters -> Score Breakdown -> Historical Evolution -> Disclaimer. No nested drilldowns, no multiple screens, no modal chains.',
         'LOCKED', 'Institutional-grade explainability kept one-hand mobile (AI Intelligence V2)'],
        ['DEC-AI-003', 'AI Intelligence',
         'No Widget Collapse on missing data: render the full widget structure and show missing values as "--"; forbidden to replace the section with "Data gathering in progress". Maintain layout stability.',
         'LOCKED', 'Layout stability and premium feel with thin data payloads (AI Intelligence V2)'],
    ]
    n_ad = 0
    for r in decisions:
        if r[0] in have:
            continue
        ad.append(r); n_ad += 1
    print('Architecture Decisions appended', n_ad)

    # Backlog: A ID,B Cat,C Epic,D Item,E Desc,F Pri,G Status,H Owner,I Branch,J Tech,K UAT,L PO,M Notes
    bl = wb['Backlog']; have_b = ids(bl)
    backlog = [
        ['CR-AI-010', 'Stock Intelligence', 'AI Intelligence V2',
         'AI Intelligence V2 implementation: KPI cards, single bottom-sheet explainability, no-collapse policy, score vs directional KPI families',
         'Implements DEC-AI-001/002/003; V1 deprecated', 'P0', 'READY FOR IMPLEMENTATION', 'HERMES', BR,
         'DESIGN LOCKED', 'PENDING', 'APPROVED', 'Spec: ' + SPEC + '; design score 10/10; PO approved'],
    ]
    n_b = 0
    for r in backlog:
        if r[0] in have_b:
            continue
        bl.append(r); n_b += 1
    print('Backlog appended', n_b)

    # CHANGELOG sheet: A Date,B Commit,C Author,D Message
    ch = wb['CHANGELOG']
    if ch.max_row == 1 and (ch['A1'].value in (None, '')):
        ch['A1'], ch['B1'], ch['C1'], ch['D1'] = 'Date', 'Commit', 'Author', 'Message'
    have_c = {str(c.value).strip() for c in ch['B'] if c.value}
    rows = [('2026-06-10', 'AI-V2-LOCK', 'ATHENA',
             'AI Intelligence V2 Design Lock: KPI cards, single bottom-sheet explainability, no-collapse policy approved; DEC-AI-001/002/003 LOCKED; CR-AI-010 READY; V1 deprecated')]
    n_ch = 0
    for d, c, a, m in rows:
        if c in have_c:
            continue
        ch.append([d, c, a, m]); n_ch += 1
    print('CHANGELOG appended', n_ch)

    wb.save(P); print('SAVED', P)


if __name__ == '__main__':
    main()
