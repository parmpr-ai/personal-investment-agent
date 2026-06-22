#!/usr/bin/env python3
"""AI Intelligence Compact V3 documentation sync into the XLSX (matches Markdown).
Backlog rows, DEC-AI-CV3 design-lock decision, CHANGELOG rows. Idempotent; backup."""
import openpyxl, shutil, datetime, os

P = os.path.join('docs', 'PIA_MASTER_BACKLOG_SOURCE_OF_TRUTH.xlsx')
BR = 'feat/pia-v3-foundation-integration'


def ids(ws, col='A'):
    return {str(c.value).strip() for c in ws[col] if c.value not in (None, '')}


def main():
    bak = P + '.bak.' + datetime.datetime.now(datetime.UTC).strftime('%Y%m%d%H%M%S')
    shutil.copy2(P, bak); print('backup', bak)
    wb = openpyxl.load_workbook(P)

    bl = wb['Backlog']; have_b = ids(bl)
    rows = [
        ['ARTEMIS-AI-COMPACT-REDESIGN-001', 'Stock Intelligence', 'AI Intelligence Compact V3',
         'AiIntelligenceCompactV3 premium compact widget: 3 rows x 4 cards, 2.2 visible cards per row; no Last Updated, no score badge, no dots/arrows',
         'Locked Compact V3 design', 'P0', 'IMPLEMENTED', 'ARTEMIS', BR, 'PASS', 'PENDING', 'PENDING', 'Commit 1b7d426'],
        ['CR-AI-COMPACT-REDESIGN-002', 'Stock Intelligence', 'AI Intelligence Compact V3',
         'Card customization: show/hide, drag reorder, three-dot Customize AI Cards sheet; persisted preferences',
         'Card source pool + customization framework', 'P1', 'IMPLEMENTED', 'ARTEMIS', BR, 'PASS', 'PENDING', 'PENDING', 'Commit 3887882'],
        ['CR-AI-COMPACT-REDESIGN-003', 'Stock Intelligence', 'AI Intelligence Compact V3',
         'Semantic tone engine: tone drives card border colour, icon glow, mini-chart stroke; Level High=red, Low=green; a BUY widget may contain red cards',
         'Semantic card coloring', 'P1', 'IMPLEMENTED', 'ARTEMIS', BR, 'PASS', 'PENDING', 'PENDING', 'Commit 3887882'],
        ['CR-AI-COMPACT-V3-UAT', 'Stock Intelligence', 'AI Intelligence Compact V3',
         'UAT PASS decision pending: NVDA BUY / NBIS HOLD / AAPL HOLD widget + customize screenshots captured (390/430)',
         'Screenshots: frontend/uat-screenshots/cr-ai-compact-v3-cr002/', 'P1', 'PENDING UAT', 'APOLLO / PO', BR, 'PASS', 'PENDING PASS DECISION', 'PENDING', 'Awaiting PO PASS decision'],
    ]
    n_b = 0
    for r in rows:
        if r[0] in have_b:
            continue
        bl.append(r); n_b += 1
    print('Backlog appended', n_b)

    ad = wb['Architecture Decisions']; have_ad = ids(ad)
    dec = ['DEC-AI-CV3', 'AI Intelligence',
           'AI Intelligence Compact V3 design lock principles: (1) no Last Updated; (2) no score badge; (3) no dots/arrows; (4) 3 rows; (5) 4 cards per row; (6) 2.2 visible cards per row; (7) card customization (show/hide, reorder, persisted); (8) semantic card coloring (tone -> border/glow/chart stroke).',
           'LOCKED', 'Approved Compact V3 redesign; commits 1b7d426, 3887882']
    n_ad = 0
    if dec[0] not in have_ad:
        ad.append(dec); n_ad += 1
    print('Architecture Decisions appended', n_ad)

    ch = wb['CHANGELOG']
    if ch.max_row == 1 and (ch['A1'].value in (None, '')):
        ch['A1'], ch['B1'], ch['C1'], ch['D1'] = 'Date', 'Commit', 'Author', 'Message'
    have_c = {str(c.value).strip() for c in ch['B'] if c.value}
    crows = [
        ('2026-06-22', '1b7d426', 'ARTEMIS', 'AiIntelligenceCompactV3 premium compact widget (ARTEMIS-AI-COMPACT-REDESIGN-001)'),
        ('2026-06-22', '3887882', 'ARTEMIS', 'Compact V3 card customization + semantic tones (CR-AI-COMPACT-REDESIGN-002/003)'),
    ]
    n_ch = 0
    for d, c, a, m in crows:
        if c not in have_c:
            ch.append([d, c, a, m]); n_ch += 1
    print('CHANGELOG appended', n_ch)

    wb.save(P); print('SAVED', P)


if __name__ == '__main__':
    main()
